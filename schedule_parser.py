from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


MONTHS_ID = {
    "januari": 1,
    "februari": 2,
    "maret": 3,
    "april": 4,
    "mei": 5,
    "juni": 6,
    "juli": 7,
    "agustus": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "desember": 12,
}

TIME_PATTERN = re.compile(r"\b([01]?\d|2[0-3])[:.]([0-5]\d)\b")
TASK_SPLIT_PATTERN = re.compile(r"\s*(?:,|/|\+|\n|;)\s*")
SPECIAL_TASK_CODES = {"duty", "off"}


@dataclass(frozen=True)
class ParsedSchedule:
    events: list[dict]
    tasks: dict[str, dict]
    month: int
    year: int


def parse_schedule_csv(
    path: str | Path,
    default_start_time: str = "08:00",
) -> ParsedSchedule:
    rows = _read_csv(path)
    month, year = _detect_month_year(rows)
    tasks = _extract_tasks(rows)
    events = _extract_events(rows, month, year, tasks, default_start_time)
    return ParsedSchedule(events=events, tasks=tasks, month=month, year=year)


def parse_schedule_file(
    path: str | Path,
    default_start_time: str = "08:00",
) -> ParsedSchedule:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return parse_schedule_csv(path, default_start_time)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return parse_schedule_xlsx(path, default_start_time)
    raise ValueError("Format file belum didukung. Gunakan .csv atau .xlsx.")


def parse_schedule_xlsx(
    path: str | Path,
    default_start_time: str = "08:00",
) -> ParsedSchedule:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = _worksheet_values(sheet)
    month, year = _detect_month_year(rows)
    tasks = _extract_xlsx_tasks(sheet)
    events = _extract_xlsx_events(sheet, month, year, tasks, default_start_time)
    return ParsedSchedule(events=events, tasks=tasks, month=month, year=year)


def _read_csv(path: str | Path) -> list[list[str]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return [[cell.strip() for cell in row] for row in csv.reader(handle)]


def _worksheet_values(sheet) -> list[list[str]]:
    rows = []
    for row in sheet.iter_rows():
        values = []
        for cell in row:
            values.append("" if cell.value is None else str(cell.value).strip())
        rows.append(values)
    return rows


def _detect_month_year(rows: list[list[str]]) -> tuple[int, int]:
    for row in rows[:10]:
        text = " ".join(cell for cell in row if cell).lower()
        for month_name, month_number in MONTHS_ID.items():
            if month_name in text:
                year_match = re.search(r"\b(20\d{2})\b", text)
                if year_match:
                    return month_number, int(year_match.group(1))
    raise ValueError("Bulan dan tahun tidak ditemukan di file schedule.")


def _extract_tasks(rows: list[list[str]]) -> dict[str, dict]:
    tasks: dict[str, dict] = {}
    for row in rows:
        if not row or not row[0]:
            continue
        code = row[0].strip()
        if _is_task_code(code):
            tasks[code] = {
                "description": _cell(row, 1) or code,
                "capacity": _last_non_empty(row[2:-1]),
                "duration": _last_non_empty(row),
            }
    return tasks


def _extract_xlsx_tasks(sheet) -> dict[str, dict]:
    tasks: dict[str, dict] = {}
    for row in sheet.iter_rows():
        for index, cell in enumerate(row):
            code = _to_text(cell.value)
            if not _is_task_code(code):
                continue
            description_cell = row[index + 1] if index + 1 < len(row) else None
            description = (
                _to_text(description_cell.value)
                if description_cell and description_cell.value is not None
                else code
            )
            tasks[code] = {
                "description": description,
                "capacity": _last_non_empty(
                    [_to_text(item.value) for item in row[index + 2:-1]]
                ),
                "duration": _last_non_empty([_to_text(item.value) for item in row]),
                "color": _cell_fill_rgb(cell),
            }
            break
    return tasks


def _extract_events(
    rows: list[list[str]],
    month: int,
    year: int,
    tasks: dict[str, dict],
    default_start_time: str,
) -> list[dict]:
    events: list[dict] = []
    row_index = 0
    while row_index < len(rows):
        row = rows[row_index]
        if _cell(row, 0).lower() != "no" or _cell(row, 1).lower() != "nama":
            row_index += 1
            continue

        date_row_index = row_index + 1
        while date_row_index < len(rows) and not _row_has_day_numbers(
            rows[date_row_index]
        ):
            date_row_index += 1
        if date_row_index >= len(rows):
            break

        day_by_col = _day_columns(rows[date_row_index])
        data_index = date_row_index + 1
        while data_index < len(rows):
            data_row = rows[data_index]
            first = _cell(data_row, 0).lower()
            second = _cell(data_row, 1)
            if first == "no" or _is_task_code(first):
                break
            if first.isdigit() and second:
                person = second
                for col, day in day_by_col.items():
                    raw_value = _cell(data_row, col)
                    if raw_value:
                        events.extend(
                            _events_from_cell(
                                person,
                                raw_value,
                                date(year, month, day),
                                tasks,
                                default_start_time,
                            )
                        )
            data_index += 1
        row_index = data_index
    return events


def _extract_xlsx_events(
    sheet,
    month: int,
    year: int,
    tasks: dict[str, dict],
    default_start_time: str,
) -> list[dict]:
    color_to_task = {
        task["color"]: code
        for code, task in tasks.items()
        if task.get("color")
    }
    events: list[dict] = []
    row_index = 1
    while row_index <= sheet.max_row:
        first = _to_text(sheet.cell(row_index, 2).value).lower()
        second = _to_text(sheet.cell(row_index, 3).value).lower()
        if first != "no" or second != "nama":
            row_index += 1
            continue

        date_row_index = row_index + 1
        while date_row_index <= sheet.max_row and not _xlsx_row_has_day_numbers(
            sheet,
            date_row_index,
        ):
            date_row_index += 1
        if date_row_index > sheet.max_row:
            break

        day_by_col = _xlsx_day_columns(sheet, date_row_index)
        data_index = date_row_index + 1
        while data_index <= sheet.max_row:
            first_cell = _to_text(sheet.cell(data_index, 2).value)
            person = _to_text(sheet.cell(data_index, 3).value)
            first_lower = first_cell.lower()
            if first_lower == "no" or _is_task_code(first_lower):
                break
            if first_cell.isdigit() and person:
                for col, day in day_by_col.items():
                    cell = sheet.cell(data_index, col)
                    raw_value = _to_text(cell.value)
                    fill_color = _cell_fill_rgb(cell)
                    task_code = _task_code_from_cell(
                        raw_value,
                        fill_color,
                        tasks,
                        color_to_task,
                    )
                    if task_code:
                        events.append(
                            _event_from_task(
                                person=person,
                                raw_value=raw_value or fill_color or task_code,
                                event_date=date(year, month, day),
                                task_code=task_code,
                                tasks=tasks,
                                default_start_time=default_start_time,
                            )
                        )
            data_index += 1
        row_index = data_index
    return events


def _events_from_cell(
    person: str,
    raw_value: str,
    event_date: date,
    tasks: dict[str, dict],
    default_start_time: str,
) -> list[dict]:
    start_time = _extract_time(raw_value)
    cleaned = TIME_PATTERN.sub("", raw_value).strip(" -")
    tokens = [token for token in TASK_SPLIT_PATTERN.split(cleaned) if token]
    if not tokens:
        tokens = [cleaned or raw_value]

    events = []
    for token in tokens:
        code = _canonical_task_code(token, tasks)
        events.append(
            _event_from_task(
                person=person,
                raw_value=raw_value,
                event_date=event_date,
                task_code=code,
                tasks=tasks,
                default_start_time=default_start_time,
                start_time=start_time,
            )
        )
    return events


def _event_from_task(
    person: str,
    raw_value: str,
    event_date: date,
    task_code: str,
    tasks: dict[str, dict],
    default_start_time: str,
    start_time: str | None = None,
) -> dict:
    start_time = start_time or _extract_time(raw_value)
    description = tasks.get(task_code, {}).get("description", task_code)
    starts_at = datetime.combine(
        event_date,
        _parse_time(start_time or default_start_time),
    ).isoformat(timespec="minutes")
    return {
        "person": person,
        "date": event_date.isoformat(),
        "start_time": start_time,
        "starts_at": starts_at,
        "task_code": task_code,
        "description": description,
        "raw_value": raw_value,
        "color": tasks.get(task_code, {}).get("color", ""),
    }


def _task_code_from_cell(
    raw_value: str,
    fill_color: str | None,
    tasks: dict[str, dict],
    color_to_task: dict[str, str],
) -> str | None:
    if raw_value:
        return _canonical_task_code(raw_value, tasks)
    if fill_color:
        return color_to_task.get(fill_color)
    return None


def _canonical_task_code(value: str, tasks: dict[str, dict]) -> str:
    normalized = re.sub(r"\s+", " ", value.strip()).lower()
    for code in tasks:
        if normalized == code.lower():
            return code

    task_number = re.search(r"\btask\s*(\d+)\b", normalized)
    if task_number:
        wanted = f"Task {task_number.group(1)}"
        for code in tasks:
            if code.lower() == wanted.lower():
                return code
        return wanted

    if normalized == "duty":
        return "Duty"
    if normalized == "off":
        return "Off"
    return value.strip()


def _is_task_code(value: str) -> bool:
    lowered = value.lower()
    return lowered.startswith("task") or lowered in SPECIAL_TASK_CODES


def _extract_time(value: str) -> str | None:
    match = TIME_PATTERN.search(value)
    if not match:
        return None
    return f"{int(match.group(1)):02d}:{match.group(2)}"


def _parse_time(value: str) -> time:
    hour, minute = value.replace(".", ":").split(":", 1)
    return time(hour=int(hour), minute=int(minute))


def _row_has_day_numbers(row: list[str]) -> bool:
    return any(cell.isdigit() and 1 <= int(cell) <= 31 for cell in row)


def _day_columns(row: list[str]) -> dict[int, int]:
    return {index: int(cell) for index, cell in enumerate(row) if cell.isdigit()}


def _xlsx_row_has_day_numbers(sheet, row_index: int) -> bool:
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row_index, col).value
        if isinstance(value, int) and 1 <= value <= 31:
            return True
        if isinstance(value, str) and value.isdigit() and 1 <= int(value) <= 31:
            return True
    return False


def _xlsx_day_columns(sheet, row_index: int) -> dict[int, int]:
    columns = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row_index, col).value
        if isinstance(value, int) and 1 <= value <= 31:
            columns[col] = value
        elif isinstance(value, str) and value.isdigit() and 1 <= int(value) <= 31:
            columns[col] = int(value)
    return columns


def _cell(row: list[str], index: int) -> str:
    return row[index].strip() if index < len(row) else ""


def _last_non_empty(row: list[str]) -> str:
    for cell in reversed(row):
        if cell:
            return cell
    return ""


def _to_text(value) -> str:
    return "" if value is None else str(value).strip()


def _cell_fill_rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fill_type:
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    return None
